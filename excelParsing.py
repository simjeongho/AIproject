import pandas as pd
import openpyxl
wb = openpyxl.load_workbook(
    "housePricePrediction.xlsx", read_only=False, data_only=False)
ws = wb['Sheet1']

# data

var = []
# 동 추출
for x in range(18, ws.max_row+1):
    item = str(ws.cell(row=x, column=1).value)
    realitem = item[-3:]
    var.append(realitem)
    ws.cell(row=x, column=2).value = realitem
# print(var)


# 공원 정보

wb.save('housePricePrediction2.xlsx')
