import json
import openpyxl
import pandas as pd
import numpy as np
import re
wb = openpyxl.load_workbook("jsontest.xlsx", read_only=False, data_only=False)
ws = wb['Sheet1']

with open("park.json", "r", encoding="utf-8") as j:
    data = json.load(j)
var = []

for x in range(1, 256):
    name = data['features'][x]['attributes']['NAM']
    addr = data['features'][x]['attributes']["ADDR"]
    ws.cell(row=x, column=1).value = name
    ws.cell(row=x, column=2).value = addr
    # print(addr)


# w 동이 들어간 곳 추출
for x in range(1, 256):
    addr = data['features'][x]['attributes']["ADDR"]
    addrToken = addr.split(" ")
    for y in range(0, len(addrToken)):
        if(addrToken[y] == "동구" or addrToken[y] == "남동구" or addrToken[y] == "인천광역시동구"):
            continue
        if(addrToken[y].find("동") != -1):
            newToken = re.sub('[^a-zA-Zㄱ-힗]', '', addrToken[y])
            loc = newToken.find("동")
            newNew = newToken[0:loc+1]
            print("newToken: " + newNew)
            ws.cell(row=x, column=3).value = newNew
            var.append(newNew)
        ws.cell(row=x, column=4+y).value = addrToken[y]
print(var)
# print(pd.Series(var).value_counts())
print(np.unique(var, return_counts=True))
count = np.unique(var, return_counts=True)
print(count[0])
wb.save('jsontest1.xlsx')
